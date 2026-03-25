"""
AIREDEX CSV → Excel Template Populator
Reads the research tool CSV export and generates one pre-populated
Excel review template per school.

Usage:
    python3 populate_templates.py AIREDEX_Research_2026-03-25.csv

Output:
    One .xlsx file per school in ./output/ folder
"""

import sys
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY     = "0D1B2A"
GOLD     = "C9A84C"
GOLD_LT  = "FFF3CC"
ICE      = "E8F0F7"
GREEN_LT = "E8F8F0"
RED_LT   = "FDE8E8"
AMBER_LT = "FEF5E0"
BLUE_LT  = "EBF5FB"
GRAY     = "F5F5F5"
WHITE    = "FFFFFF"
MID_GRAY = "CCCCCC"
GREEN_VERIFIED = "D5F5E3"
AMBER_INFERRED = "FEF9E7"
RED_ASSUMED    = "FADBD8"

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(size=10, bold=False, color="1A2B3C", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def border():
    s = Side(style='thin', color=MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def conf_color(conf):
    conf = (conf or "").lower()
    if conf == "verified":  return GREEN_VERIFIED
    if conf == "inferred":  return AMBER_INFERRED
    if conf == "assumed":   return RED_ASSUMED
    return WHITE

def conf_label(conf):
    conf = (conf or "").lower()
    if conf == "verified":  return "✓ Verified"
    if conf == "inferred":  return "~ Inferred"
    if conf == "assumed":   return "✗ Assumed — REVIEW"
    return conf

# ── Pillar definitions ─────────────────────────────────────────────────────────
PILLARS = {
    "p1": {
        "name": "Institutional Commitment",
        "scope": "Institution-wide",
        "max_r": 27, "max_u": 30,
        "conf_key": "p1_conf",
    },
    "p2": {
        "name": "Curriculum Integration",
        "scope": "UNDERGRADUATE ONLY",
        "max_r": 30, "max_u": 40,
        "conf_key": "p2_conf",
    },
    "p3": {
        "name": "Industry Partnerships",
        "scope": "Institution-wide",
        "max_r": 20, "max_u": 12,
        "conf_key": "p3_conf",
    },
    "p4": {
        "name": "Research Depth & Undergrad Access",
        "scope": "Institution-wide (undergrad access required)",
        "max_r": 16, "max_u": 8,
        "conf_key": "p4_conf",
    },
    "p5": {
        "name": "A(I)ttitude",
        "scope": "Institution-wide — Editorial judgment",
        "max_r": 7, "max_u": 10,
        "conf_key": "p5_conf",
    },
}

TIER_LABELS = {
    "S": "Exceptional AI Readiness",
    "A": "Strong AI Readiness",
    "B": "Developing AI Readiness",
    "C": "Emerging AI Readiness",
    "D": "Limited AI Readiness",
}

def get_tier(score, track):
    score = int(score or 0)
    if track == "R":
        if score >= 90: return "S"
        if score >= 75: return "A"
        if score >= 60: return "B"
        if score >= 45: return "C"
        return "D"
    else:
        if score >= 80: return "S"
        if score >= 63: return "A"
        if score >= 48: return "B"
        if score >= 33: return "C"
        return "D"

def set_cell(ws, row, col, value, fnt=None, fll=None, aln=None, brd=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fll: c.fill = fll
    if aln: c.alignment = aln
    if brd: c.border = brd
    return c

# ── Build one Excel per school ─────────────────────────────────────────────────
def build_template(school, output_dir):
    name    = school.get("University", "Unknown")
    state   = school.get("State", "")
    track   = school.get("Track", "R") if "Track" in school else "R"
    cat     = school.get("Category", "National University") if "Category" in school else "National University"
    rank    = school.get("AI Rank", "")
    total   = school.get("Total Score", "0")
    notes   = school.get("Research Notes", "")
    sources = school.get("Key Evidence", "")
    alumni  = school.get("Notable Alumni", "")
    search_date = school.get("Search Date", "")
    freshness   = school.get("Freshness", "")
    newest_src  = school.get("Newest Source", "")
    posture     = school.get("A(I)ttitude", school.get("Posture", ""))

    tier = get_tier(total, track)

    wb = openpyxl.Workbook()

    # ── SHEET 1: OVERVIEW ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "📊 Overview"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 55
    ws.row_dimensions[1].height = 8

    # Title bar
    ws.row_dimensions[2].height = 36
    ws.merge_cells('B2:C2')
    set_cell(ws, 2, 2, f"AIREDEX™  —  {name}",
        fnt=Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        fll=fill(NAVY),
        aln=Alignment(horizontal="center", vertical="center"))

    ws.row_dimensions[3].height = 18
    ws.merge_cells('B3:C3')
    freshness_flag = {"green": "🟢", "yellow": "🟡", "red": "🔴"}.get(freshness.lower(), "")
    set_cell(ws, 3, 2,
        f"AI Readiness Index  |  {cat}  |  {track} Track  |  Research: {search_date}  {freshness_flag}",
        fnt=Font(name="Arial", size=9, italic=True, color=GOLD),
        fll=fill(NAVY),
        aln=Alignment(horizontal="center", vertical="center"))

    ws.row_dimensions[4].height = 10

    # Score summary box
    r = 5
    fields = [
        ("Institution", name),
        ("State", state),
        ("Category", cat),
        ("AIREDEX Track", f"{track} Track ({'Research University' if track=='R' else 'Undergraduate College'})"),
        ("Application Rank (2025)", rank),
        ("Total Score", f"{total}/{'100' if track=='R' else '100'}"),
        ("AIREDEX Tier", f"{tier}  —  {TIER_LABELS.get(tier, '')}"),
        ("A(I)ttitude / Posture", posture),
        ("Research Date", search_date),
        ("Newest Source", newest_src),
        ("Freshness", freshness.upper() if freshness else ""),
    ]

    for label, value in fields:
        ws.row_dimensions[r].height = 28
        is_score = label == "Total Score"
        is_tier  = label == "AIREDEX Tier"
        tier_colors = {"S":"FFF3CC","A":"D5F5E3","B":"D6EAF8","C":"FEF5E0","D":"FADBD8"}
        row_fill = fill(tier_colors.get(tier, WHITE)) if is_tier else (fill(GOLD_LT) if is_score else fill(GRAY if r%2==0 else WHITE))

        set_cell(ws, r, 2, label,
            fnt=font(10, bold=True),
            fll=row_fill, aln=align(), brd=border())
        set_cell(ws, r, 3, value,
            fnt=font(11 if is_score or is_tier else 10, bold=is_score or is_tier),
            fll=row_fill, aln=align(), brd=border())
        r += 1

    r += 1
    # Pillar score summary
    ws.row_dimensions[r].height = 24
    for col, hdr in zip([2,3,4,5,6], ["Pillar","Score","Max","Confidence","Status"]):
        ws.column_dimensions[get_column_letter(col)].width = [30,10,10,16,20][col-2]
        set_cell(ws, r, col, hdr,
            fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            fll=fill(NAVY), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())

    for pk, pdef in PILLARS.items():
        r += 1
        ws.row_dimensions[r].height = 28
        score_val = school.get(f"{pk.upper()} {pdef['name'].split()[0]}", school.get(pk.upper(), "0"))
        # Try multiple CSV column name formats
        for key in [pk, pk.upper(), f"P{pk[1]} {pdef['name'].split()[0]}", f"p{pk[1]}"]:
            if key in school:
                score_val = school[key]
                break
        
        conf_val = school.get(pdef["conf_key"], school.get(f"{pk}_conf", ""))
        max_val = pdef["max_r"] if track == "R" else pdef["max_u"]
        c_color = conf_color(conf_val)
        needs_review = conf_val.lower() in ["assumed", "inferred"] if conf_val else True

        set_cell(ws, r, 2, f"P{pk[1]} — {pdef['name']}",
            fnt=font(10), fll=fill(c_color), aln=align(), brd=border())
        set_cell(ws, r, 3, score_val,
            fnt=font(11, bold=True, color=NAVY),
            fll=fill(c_color), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())
        set_cell(ws, r, 4, max_val,
            fnt=font(10, color="888888"),
            fll=fill(c_color), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())
        set_cell(ws, r, 5, conf_label(conf_val),
            fnt=font(10, bold=needs_review),
            fll=fill(c_color), aln=align("center"), brd=border())
        set_cell(ws, r, 6,
            "⚠ REVIEW NEEDED" if conf_val.lower() == "assumed" else
            ("~ Verify source" if conf_val.lower() == "inferred" else "✓ OK to approve"),
            fnt=font(10, bold=needs_review,
                color="C0392B" if conf_val.lower()=="assumed" else
                      "8B6914" if conf_val.lower()=="inferred" else "1E8449"),
            fll=fill(c_color), aln=align(), brd=border())

    r += 2
    # Research notes
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "AI RESEARCH NOTES",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    r += 1
    ws.row_dimensions[r].height = 80
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, notes,
        fnt=font(10), fll=fill(WHITE), aln=align(), brd=border())

    r += 1
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "KEY EVIDENCE",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    r += 1
    ws.row_dimensions[r].height = 60
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, sources,
        fnt=font(10), fll=fill(WHITE), aln=align(), brd=border())

    r += 2
    # Supervisor review section
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "SUPERVISOR REVIEW",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    review_items = [
        "Scores reviewed and verified against sources",
        "Manual criteria completed (P2 depth, cross-disciplinary count)",
        "Commitment Bonus applied if warranted (max +5)",
        "Editorial flags reviewed",
        "Approved for database entry — Supervisor initials + date:",
    ]
    for item in review_items:
        r += 1
        ws.row_dimensions[r].height = 28
        set_cell(ws, r, 2, item,
            fnt=font(10), fll=fill(ICE), aln=align(), brd=border())
        ws.merge_cells(f'C{r}:F{r}')
        set_cell(ws, r, 3, "",
            fll=fill(WHITE), aln=align(), brd=border())

    # ── SHEET 2: SCORE DETAIL ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("🔍 Score Detail")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 2
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 14

    ws2.row_dimensions[1].height = 8
    ws2.row_dimensions[2].height = 32
    ws2.merge_cells('B2:D2')
    set_cell(ws2, 2, 2, f"SCORE DETAIL  —  {name}",
        fnt=Font(name="Arial", size=12, bold=True, color="FFFFFF"),
        fll=fill(NAVY), aln=Alignment(horizontal="center", vertical="center"))

    r2 = 4
    for pk, pdef in PILLARS.items():
        conf_val = school.get(pdef["conf_key"], school.get(f"{pk}_conf", ""))
        score_val = "0"
        for key in [pk, pk.upper(), f"p{pk[1]}"]:
            if key in school:
                score_val = school[key]
                break

        max_val = pdef["max_r"] if track == "R" else pdef["max_u"]
        c_color = conf_color(conf_val)
        needs_review = conf_val.lower() in ["assumed", "inferred"] if conf_val else True

        ws2.row_dimensions[r2].height = 24
        ws2.merge_cells(f'B{r2}:D{r2}')
        set_cell(ws2, r2, 2,
            f"P{pk[1]} — {pdef['name'].upper()}  |  {pdef['scope']}  |  Max: {max_val} pts",
            fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            fll=fill(NAVY), aln=Alignment(horizontal="left", vertical="center"))
        r2 += 1

        rows_detail = [
            ("AI Score", score_val),
            ("Max Points", str(max_val)),
            ("Confidence", conf_label(conf_val)),
            ("Status", "⚠ REVIEW NEEDED — override score if evidence insufficient" if needs_review else "✓ Verified — spot check source URL"),
            ("Override Score", "Enter here if you disagree with AI score:"),
            ("Override Reason", ""),
        ]

        for label, val in rows_detail:
            ws2.row_dimensions[r2].height = 28
            set_cell(ws2, r2, 2, label,
                fnt=font(10, bold=True),
                fll=fill(c_color if label in ["Confidence","Status"] else GRAY),
                aln=align(), brd=border())
            set_cell(ws2, r2, 3, val,
                fnt=font(10, bold=label=="Status", color="C0392B" if needs_review and label=="Status" else "1A2B3C"),
                fll=fill(c_color if label in ["Confidence","Status"] else WHITE),
                aln=align(), brd=border())
            ws2.merge_cells(f'C{r2}:D{r2}')
            r2 += 1

        r2 += 1

    # ── SAVE ──────────────────────────────────────────────────────────────────
    safe_name = name.replace("/", "-").replace("\\", "-").replace(":", "").replace("&", "and")
    filename = os.path.join(output_dir, f"AIREDEX_{safe_name}_{track}Track.xlsx")
    wb.save(filename)
    return filename


# ── MAIN ───────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 populate_templates.py <csv_file>")
        print("Example: python3 populate_templates.py AIREDEX_Research_2026-03-25.csv")
        sys.exit(1)

    csv_file = sys.argv[1]
    if not os.path.exists(csv_file):
        print(f"Error: File not found: {csv_file}")
        sys.exit(1)

    output_dir = "AIREDEX_Review_Templates"
    os.makedirs(output_dir, exist_ok=True)

    with open(csv_file, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        schools = list(reader)

    print(f"\nAIREDEX Template Generator")
    print(f"{'─'*40}")
    print(f"Input:   {csv_file}")
    print(f"Schools: {len(schools)}")
    print(f"Output:  ./{output_dir}/")
    print(f"{'─'*40}")

    generated = []
    for school in schools:
        name = school.get("University", "Unknown")
        try:
            filename = build_template(school, output_dir)
            generated.append(filename)
            conf_flags = []
            for pk in ["p1","p2","p3","p4","p5"]:
                conf = school.get(f"{pk}_conf", "").lower()
                if conf == "assumed":   conf_flags.append(f"P{pk[1]}⚠")
                elif conf == "inferred": conf_flags.append(f"P{pk[1]}~")
            flag_str = f"  [{', '.join(conf_flags)}]" if conf_flags else "  [all verified]"
            print(f"✓  {name}{flag_str}")
        except Exception as e:
            print(f"✗  {name} — ERROR: {e}")

    print(f"{'─'*40}")
    print(f"Generated {len(generated)} templates in ./{output_dir}/")
    print(f"\nReview priority: files with ⚠ (Assumed) need attention first.")
    print(f"Files with ~ (Inferred) should have sources spot-checked.")
    print(f"Files with all verified can be approved quickly.\n")


if __name__ == "__main__":
    main()
