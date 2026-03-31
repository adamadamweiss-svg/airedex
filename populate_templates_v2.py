"""
AIREDEX CSV to Excel Template Populator v2
Reads the research tool CSV export and generates one pre-populated
Excel review template per school.

Tracks:
    R  = National University (>20k undergrad, research-focused)
    UL = Mid-size Regional University (10k-20k undergrad)
    US = Small Regional University (<10k undergrad)

Usage:
    python3 populate_templates_v2.py AIREDEX_Research_2026-03-30.csv

Output:
    One .xlsx file per school in ./AIREDEX_Review_Templates/ folder
"""

import sys
import csv
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Colors ────────────────────────────────────────────────────────────────────
NAVY        = "0D1B2A"
GOLD        = "C9A84C"
GOLD_LT     = "FFF3CC"
ICE         = "E8F0F7"
GREEN_LT    = "E8F8F0"
RED_LT      = "FDE8E8"
AMBER_LT    = "FEF5E0"
AMBER_DARK  = "F39C12"
BLUE_LT     = "EBF5FB"
GRAY        = "F5F5F5"
WHITE       = "FFFFFF"
MID_GRAY    = "CCCCCC"
GREEN_VERIFIED = "D5F5E3"
AMBER_INFERRED = "FEF9E7"
RED_ASSUMED    = "FADBD8"

def fill(color):
    return PatternFill("solid", fgColor=color)

def font(size=10, bold=False, color="1A2B3C", italic=False):
    return Font(name="Arial", size=size, bold=bold, color=color, italic=italic)

def border():
    s = Side(style='thin', color=MID_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def conf_color(conf):
    conf = (conf or "").lower()
    if conf == "verified":  return GREEN_VERIFIED
    if conf == "inferred":  return AMBER_INFERRED
    if conf == "assumed":   return RED_ASSUMED
    return WHITE

def conf_label(conf):
    conf = (conf or "").lower()
    if conf == "verified":  return "Verified"
    if conf == "inferred":  return "~ Inferred"
    if conf == "assumed":   return "Assumed — REVIEW"
    return conf

# ── Track-aware pillar definitions ────────────────────────────────────────────
# P5 is an editorial label (enabling/progressive/neutral/cautious/resistant)
# and is not scored. max values reflect the locked rubrics.

PILLAR_DEFS = {
    "p1": {
        "name": "Institutional Commitment",
        "scope": "Institution-wide",
        "max": {"R": 40, "UL": 35, "US": 35},
        "conf_key": "P1 Conf",
        "score_key": "P1 Commitment",
    },
    "p2": {
        "name": "Curriculum Integration",
        "scope": "Undergraduate focus",
        "max": {"R": 35, "UL": 30, "US": 35},
        "conf_key": "P2 Conf",
        "score_key": "P2 Curriculum",
    },
    "p3": {
        "name": "Industry Partnerships",
        "scope": "Institution-wide",
        "max": {"R": 13, "UL": 20, "US": 10},
        "conf_key": "P3 Conf",
        "score_key": "P3 Partnerships",
    },
    "p4": {
        "name": "Research Output",
        "scope": "Institution-wide",
        "max": {"R": 12, "UL": 15, "US": 15},
        "conf_key": "P4 Conf",
        "score_key": "P4 Research",
    },
}

TIER_LABELS = {
    "S": "Exceptional AI Readiness",
    "A": "Strong AI Readiness",
    "B": "Developing AI Readiness",
    "C": "Emerging AI Readiness",
    "D": "Limited AI Readiness",
}

TRACK_LABELS = {
    "R":  "National University (R)",
    "UL": "Mid-size Regional University (U-L)",
    "US": "Small Regional University (U-S)",
}

P5_LABELS = ["enabling", "progressive", "neutral", "cautious", "resistant"]

def normalize_track(track_raw):
    """Normalize track string from CSV to R / UL / US."""
    t = (track_raw or "").strip().upper().replace("-", "").replace("_", "")
    if t in ("UL", "U-L", "UL"):
        return "UL"
    if t in ("US", "U-S", "US"):
        return "US"
    if t == "R":
        return "R"
    # Fallback: treat unknown as UL (mid-size regional)
    return "UL"

def get_tier(score, track):
    score = int(score or 0)
    if track == "R":
        if score >= 90: return "S"
        if score >= 75: return "A"
        if score >= 60: return "B"
        if score >= 45: return "C"
        return "D"
    else:
        # UL and US share same tier thresholds
        if score >= 80: return "S"
        if score >= 63: return "A"
        if score >= 48: return "B"
        if score >= 33: return "C"
        return "D"

def set_cell(ws, row, col, value, fnt=None, fll=None, aln=None, brd=None):
    c = ws.cell(row=row, column=col, value=value)
    if fnt: c.font = fnt
    if fll: c.fill = fll
    if aln: c.alignment = aln
    if brd: c.border = brd
    return c

def is_alert_flag(flag_text):
    """Return True if the flag contains keywords requiring bold-red styling."""
    lower = flag_text.lower()
    return any(kw in lower for kw in ["zero", "verify", "review"])

# ── Build one Excel per school ────────────────────────────────────────────────
def build_template(school, output_dir):
    name        = school.get("University", "Unknown")
    state       = school.get("State", "")
    track_raw   = school.get("Track", "UL")
    track       = normalize_track(track_raw)
    rank        = school.get("AI Rank", "")
    total       = school.get("Total Score", "0")
    notes       = school.get("Research Notes", "")
    key_ev      = school.get("Key Evidence", "")
    alumni      = school.get("Notable Alumni", "")
    search_date = school.get("Search Date", "")
    freshness   = school.get("Freshness", "")
    newest_src  = school.get("Newest Source", "")
    p5_label    = school.get("P5 A(I)ttitude", "").strip() or "neutral"
    review_flags_raw = school.get("Review Flags", "")

    # Parse review flags: pipe-delimited string -> list
    review_flags = [f.strip() for f in review_flags_raw.split("|") if f.strip()] if review_flags_raw else []

    tier = get_tier(total, track)
    track_display = TRACK_LABELS.get(track, track)

    wb = openpyxl.Workbook()

    # ── SHEET 1: OVERVIEW ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 55
    ws.row_dimensions[1].height = 8

    # Title bar
    ws.row_dimensions[2].height = 36
    ws.merge_cells('B2:C2')
    set_cell(ws, 2, 2, f"AIREDEX  --  {name}",
        fnt=Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        fll=fill(NAVY),
        aln=Alignment(horizontal="center", vertical="center"))

    ws.row_dimensions[3].height = 18
    ws.merge_cells('B3:C3')
    freshness_flag = {"green": "[GREEN]", "yellow": "[YELLOW]", "red": "[RED]"}.get((freshness or "").lower(), "")
    set_cell(ws, 3, 2,
        f"AI Readiness Index  |  {track_display}  |  Research: {search_date}  {freshness_flag}",
        fnt=Font(name="Arial", size=9, italic=True, color=GOLD),
        fll=fill(NAVY),
        aln=Alignment(horizontal="center", vertical="center"))

    ws.row_dimensions[4].height = 10

    # Score summary box
    r = 5
    fields = [
        ("Institution",             name),
        ("State",                   state),
        ("AIREDEX Track",           track_display),
        ("Application Rank",        rank),
        ("Total Score",             f"{total} / 100"),
        ("AIREDEX Tier",            f"{tier}  --  {TIER_LABELS.get(tier, '')}"),
        ("P5 A(I)ttitude",          p5_label.upper()),
        ("Research Date",           search_date),
        ("Newest Source",           newest_src),
        ("Freshness",               (freshness or "").upper()),
    ]

    for label, value in fields:
        ws.row_dimensions[r].height = 28
        is_score = label == "Total Score"
        is_tier  = label == "AIREDEX Tier"
        is_p5    = label == "P5 A(I)ttitude"
        tier_colors = {"S": "FFF3CC", "A": "D5F5E3", "B": "D6EAF8", "C": "FEF5E0", "D": "FADBD8"}
        if is_tier:
            row_fill = fill(tier_colors.get(tier, WHITE))
        elif is_score:
            row_fill = fill(GOLD_LT)
        elif is_p5:
            row_fill = fill(AMBER_LT)
        else:
            row_fill = fill(GRAY if r % 2 == 0 else WHITE)

        set_cell(ws, r, 2, label,
            fnt=font(10, bold=True),
            fll=row_fill, aln=align(), brd=border())
        set_cell(ws, r, 3, value,
            fnt=font(11 if is_score or is_tier else 10, bold=is_score or is_tier),
            fll=row_fill, aln=align(), brd=border())
        r += 1

    r += 1

    # Pillar score summary header
    ws.row_dimensions[r].height = 24
    for col, hdr in zip([2, 3, 4, 5, 6], ["Pillar", "Score", "Max", "Confidence", "Status"]):
        ws.column_dimensions[get_column_letter(col)].width = [30, 10, 10, 16, 20][col - 2]
        set_cell(ws, r, col, hdr,
            fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            fll=fill(NAVY), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())

    for pk, pdef in PILLAR_DEFS.items():
        r += 1
        ws.row_dimensions[r].height = 28

        # Resolve score from CSV — try multiple column name patterns
        score_val = "0"
        for key in [pdef["score_key"], pk, pk.upper(), f"p{pk[1]}"]:
            if key in school and school[key] not in ("", None):
                score_val = school[key]
                break

        conf_val = school.get(pdef["conf_key"], school.get(f"{pk}_conf", "")).strip()
        max_val  = pdef["max"].get(track, 0)
        c_color  = conf_color(conf_val)
        needs_review = (conf_val.lower() in ["assumed", "inferred"]) if conf_val else True

        set_cell(ws, r, 2, f"P{pk[1]} -- {pdef['name']}",
            fnt=font(10), fll=fill(c_color), aln=align(), brd=border())
        set_cell(ws, r, 3, score_val,
            fnt=font(11, bold=True, color=NAVY),
            fll=fill(c_color), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())
        set_cell(ws, r, 4, max_val,
            fnt=font(10, color="888888"),
            fll=fill(c_color), aln=Alignment(horizontal="center", vertical="center"),
            brd=border())
        set_cell(ws, r, 5, conf_label(conf_val),
            fnt=font(10, bold=needs_review),
            fll=fill(c_color), aln=align("center"), brd=border())
        set_cell(ws, r, 6,
            "REVIEW NEEDED" if conf_val.lower() == "assumed" else
            ("~ Verify source" if conf_val.lower() == "inferred" else "OK to approve"),
            fnt=font(10, bold=needs_review,
                color="C0392B" if conf_val.lower() == "assumed" else
                      "8B6914" if conf_val.lower() == "inferred" else "1E8449"),
            fll=fill(c_color), aln=align(), brd=border())

    # P5 row — label only, no max points
    r += 1
    ws.row_dimensions[r].height = 28
    p5_conf_val = school.get("P5 Conf", school.get("p5_conf", "")).strip()
    p5_c_color  = conf_color(p5_conf_val)
    set_cell(ws, r, 2, "P5 -- A(I)ttitude (label only)",
        fnt=font(10, italic=True), fll=fill(AMBER_LT), aln=align(), brd=border())
    set_cell(ws, r, 3, p5_label.upper(),
        fnt=font(10, bold=True, color=NAVY),
        fll=fill(AMBER_LT), aln=Alignment(horizontal="center", vertical="center"),
        brd=border())
    set_cell(ws, r, 4, "label",
        fnt=font(10, color="888888", italic=True),
        fll=fill(AMBER_LT), aln=Alignment(horizontal="center", vertical="center"),
        brd=border())
    set_cell(ws, r, 5, conf_label(p5_conf_val),
        fnt=font(10), fll=fill(AMBER_LT), aln=align("center"), brd=border())
    set_cell(ws, r, 6, "Editorial judgment",
        fnt=font(10, italic=True, color="8B6914"),
        fll=fill(AMBER_LT), aln=align(), brd=border())

    r += 2

    # Research notes section
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "AI RESEARCH NOTES",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    r += 1
    ws.row_dimensions[r].height = 80
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, notes,
        fnt=font(10), fll=fill(WHITE), aln=align(), brd=border())

    r += 1
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "KEY EVIDENCE",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    r += 1
    ws.row_dimensions[r].height = 60
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, key_ev,
        fnt=font(10), fll=fill(WHITE), aln=align(), brd=border())

    r += 1

    # ── REVIEW FLAGS SECTION (only if flags exist) ────────────────────────────
    if review_flags:
        r += 1
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f'B{r}:F{r}')
        set_cell(ws, r, 2, "REVIEW FLAGS",
            fnt=Font(name="Arial", size=9, bold=True, color="C0392B"),
            fll=fill(AMBER_LT), aln=Alignment(horizontal="left", vertical="center"))

        for flag in review_flags:
            r += 1
            ws.row_dimensions[r].height = 28
            ws.merge_cells(f'B{r}:F{r}')
            alert = is_alert_flag(flag)
            set_cell(ws, r, 2, flag,
                fnt=Font(name="Arial", size=10,
                         bold=alert,
                         color="C0392B" if alert else "8B4513"),
                fll=fill(AMBER_LT), aln=align(), brd=border())

        r += 1

    # ── SUPERVISOR REVIEW SECTION ─────────────────────────────────────────────
    r += 1
    ws.row_dimensions[r].height = 20
    ws.merge_cells(f'B{r}:F{r}')
    set_cell(ws, r, 2, "SUPERVISOR REVIEW",
        fnt=Font(name="Arial", size=9, bold=True, color=GOLD),
        fll=fill(GRAY), aln=Alignment(horizontal="left", vertical="center"))

    review_items = [
        "Scores reviewed and verified against sources",
        "Manual criteria completed (P2 depth, cross-disciplinary count)",
        "Track confirmed correct (R / U-L / U-S) against enrollment data",
        "P5 A(I)ttitude label reviewed and confirmed",
        "Review Flags addressed and resolved",
        "Approved for database entry -- Supervisor initials + date:",
    ]
    for item in review_items:
        r += 1
        ws.row_dimensions[r].height = 28
        set_cell(ws, r, 2, item,
            fnt=font(10), fll=fill(ICE), aln=align(), brd=border())
        ws.merge_cells(f'C{r}:F{r}')
        set_cell(ws, r, 3, "",
            fll=fill(WHITE), aln=align(), brd=border())

    # ── SHEET 2: SCORE DETAIL ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Score Detail")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 2
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 14

    ws2.row_dimensions[1].height = 8
    ws2.row_dimensions[2].height = 32
    ws2.merge_cells('B2:D2')
    set_cell(ws2, 2, 2, f"SCORE DETAIL  --  {name}  ({track_display})",
        fnt=Font(name="Arial", size=12, bold=True, color="FFFFFF"),
        fll=fill(NAVY), aln=Alignment(horizontal="center", vertical="center"))

    r2 = 4
    for pk, pdef in PILLAR_DEFS.items():
        conf_val  = school.get(pdef["conf_key"], school.get(f"{pk}_conf", "")).strip()
        score_val = "0"
        for key in [pdef["score_key"], pk, pk.upper(), f"p{pk[1]}"]:
            if key in school and school[key] not in ("", None):
                score_val = school[key]
                break

        max_val      = pdef["max"].get(track, 0)
        c_color      = conf_color(conf_val)
        needs_review = (conf_val.lower() in ["assumed", "inferred"]) if conf_val else True

        ws2.row_dimensions[r2].height = 24
        ws2.merge_cells(f'B{r2}:D{r2}')
        set_cell(ws2, r2, 2,
            f"P{pk[1]} -- {pdef['name'].upper()}  |  {pdef['scope']}  |  Max: {max_val} pts",
            fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
            fll=fill(NAVY), aln=Alignment(horizontal="left", vertical="center"))
        r2 += 1

        rows_detail = [
            ("AI Score",       score_val),
            ("Max Points",     str(max_val)),
            ("Confidence",     conf_label(conf_val)),
            ("Status",         "REVIEW NEEDED -- override score if evidence insufficient" if needs_review else "Verified -- spot check source URL"),
            ("Override Score", "Enter here if you disagree with AI score:"),
            ("Override Reason", ""),
        ]

        for label, val in rows_detail:
            ws2.row_dimensions[r2].height = 28
            set_cell(ws2, r2, 2, label,
                fnt=font(10, bold=True),
                fll=fill(c_color if label in ["Confidence", "Status"] else GRAY),
                aln=align(), brd=border())
            set_cell(ws2, r2, 3, val,
                fnt=font(10, bold=label == "Status",
                         color="C0392B" if needs_review and label == "Status" else "1A2B3C"),
                fll=fill(c_color if label in ["Confidence", "Status"] else WHITE),
                aln=align(), brd=border())
            ws2.merge_cells(f'C{r2}:D{r2}')
            r2 += 1

        r2 += 1

    # P5 detail on sheet 2
    ws2.row_dimensions[r2].height = 24
    ws2.merge_cells(f'B{r2}:D{r2}')
    set_cell(ws2, r2, 2, "P5 -- A(I)TTITUDE  |  Editorial label only  |  Not scored",
        fnt=Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        fll=fill(NAVY), aln=Alignment(horizontal="left", vertical="center"))
    r2 += 1

    p5_detail = [
        ("Label",        p5_label.upper()),
        ("Confidence",   conf_label(school.get("P5 Conf", school.get("p5_conf", "")).strip())),
        ("Scale",        "leading edge / enabling / neutral / cautious / restrictive"),
        ("Override",     "Enter corrected label if needed:"),
    ]
    for label, val in p5_detail:
        ws2.row_dimensions[r2].height = 28
        set_cell(ws2, r2, 2, label,
            fnt=font(10, bold=True), fll=fill(AMBER_LT), aln=align(), brd=border())
        set_cell(ws2, r2, 3, val,
            fnt=font(10), fll=fill(AMBER_LT if label in ["Label", "Confidence"] else WHITE),
            aln=align(), brd=border())
        ws2.merge_cells(f'C{r2}:D{r2}')
        r2 += 1

    # ── SAVE ──────────────────────────────────────────────────────────────────
    safe_name = (name.replace("/", "-").replace("\\", "-")
                     .replace(":", "").replace("&", "and"))
    filename = os.path.join(output_dir, f"AIREDEX_{safe_name}_{track}Track.xlsx")
    wb.save(filename)
    return filename


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("Usage: python3 populate_templates_v2.py <csv_file>")
        print("Example: python3 populate_templates_v2.py airedex_utrack_calibration.csv")
        sys.exit(1)

    csv_file = sys.argv[1]
    if not os.path.exists(csv_file):
        print(f"Error: File not found: {csv_file}")
        sys.exit(1)

    output_dir = "AIREDEX_Review_Templates"
    os.makedirs(output_dir, exist_ok=True)

    # Skip comment rows (lines starting with #) before passing to DictReader
    with open(csv_file, newline='', encoding='utf-8') as f:
        lines = [l for l in f if not l.startswith('#')]

    reader = csv.DictReader(lines)
    schools = list(reader)

    print(f"\nAIREDEX Template Generator v2")
    print(f"{'─' * 44}")
    print(f"Input:   {csv_file}")
    print(f"Schools: {len(schools)}")
    print(f"Output:  ./{output_dir}/")
    print(f"{'─' * 44}")

    generated = []
    for school in schools:
        name = school.get("University", "Unknown")
        # Skip comment or blank rows that slipped through
        if not name or name.startswith("#"):
            continue
        try:
            filename = build_template(school, output_dir)
            generated.append(filename)

            track = normalize_track(school.get("Track", "UL"))
            conf_flags = []
            for pk in ["p1", "p2", "p3", "p4", "p5"]:
                conf_key = f"P{pk[1]} Conf" if pk != "p5" else "P5 Conf"
                conf = school.get(conf_key, school.get(f"{pk}_conf", "")).lower()
                if conf == "assumed":    conf_flags.append(f"P{pk[1]} ASSUMED")
                elif conf == "inferred": conf_flags.append(f"P{pk[1]}~")

            review_flags_raw = school.get("Review Flags", "")
            flag_count = len([f for f in review_flags_raw.split("|") if f.strip()]) if review_flags_raw else 0
            flag_note = f"  [{flag_count} review flag{'s' if flag_count != 1 else ''}]" if flag_count else ""

            conf_str = f"  [{', '.join(conf_flags)}]" if conf_flags else "  [all verified]"
            print(f"OK  {name} ({track}){conf_str}{flag_note}")

        except Exception as e:
            print(f"ERR {name} -- {e}")

    print(f"{'─' * 44}")
    print(f"Generated {len(generated)} templates in ./{output_dir}/")
    print(f"\nReview priority:")
    print(f"  ASSUMED  -- must override before approving")
    print(f"  ~        -- verify source URL")
    print(f"  flags    -- editorial review flags require human judgment\n")


if __name__ == "__main__":
    main()
